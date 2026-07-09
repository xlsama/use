#!/usr/bin/env python3
"""
Excel to Markdown Converter

Supported formats:
    .xlsx   Excel workbook
    .xlsm   Excel macro-enabled workbook

Unsupported by default:
    .xls    Legacy binary Excel format; resave as .xlsx first

All paths produce the same output convention:
    <input>.md                     Markdown file
"""

import argparse
import re
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any


# ─────────────────────────────────────────────────────────────
# Format registry
# ─────────────────────────────────────────────────────────────

EXCEL_FORMATS = {".xlsx", ".xlsm"}
LEGACY_EXCEL_FORMATS = {".xls"}


# ─────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────

def _format_size(size: int) -> str:
    for unit in ("B", "KB", "MB"):
        if size < 1024:
            return f"{size:.0f} {unit}"
        size /= 1024
    return f"{size:.1f} GB"


def _report_result(out_file: Path) -> None:
    size = out_file.stat().st_size
    print(f"[OK] Saved Markdown to: {out_file} ({_format_size(size)})")


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _markdown_escape(value: str) -> str:
    value = value.replace("\\", "\\\\")
    value = value.replace("|", "\\|")
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"\s*\n\s*", "<br>", value).strip()


def _format_cell_value(value: Any) -> str:
    if _is_empty(value):
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, float):
        return _markdown_escape(f"{value:g}")
    return _markdown_escape(str(value))


def _is_numeric_value(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def _sheet_state_label(sheet_state: str) -> str:
    if sheet_state == "visible":
        return "visible"
    return sheet_state or "unknown"


# ─────────────────────────────────────────────────────────────
# Worksheet extraction
# ─────────────────────────────────────────────────────────────

def _merged_value_map(worksheet) -> dict[tuple[int, int], Any]:
    """Return propagated values for merged cells, keyed by (row, column).

    Merged regions whose top-left cell is empty are intentionally skipped.
    These are typically formatting-only ranges that carry no textual content.
    """
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in worksheet.merged_cells.ranges:
        value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        if _is_empty(value):
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = value
    return merged_values


def _cell_value(worksheet, row: int, col: int, merged_values: dict[tuple[int, int], Any]) -> Any:
    value = worksheet.cell(row, col).value
    if _is_empty(value):
        return merged_values.get((row, col), value)
    return value


def _content_bounds(worksheet, merged_values: dict[tuple[int, int], Any]) -> tuple[int, int, int, int] | None:
    min_row = min_col = None
    max_row = max_col = None

    for row in worksheet.iter_rows():
        for cell in row:
            if _is_empty(cell.value):
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            max_row = cell.row if max_row is None else max(max_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_col = cell.column if max_col is None else max(max_col, cell.column)

    for row, col in merged_values:
        min_row = row if min_row is None else min(min_row, row)
        max_row = row if max_row is None else max(max_row, row)
        min_col = col if min_col is None else min(min_col, col)
        max_col = col if max_col is None else max(max_col, col)

    if min_row is None or min_col is None or max_row is None or max_col is None:
        return None
    return min_row, min_col, max_row, max_col


def _trim_trailing_empty_cells(row: list[Any]) -> list[Any]:
    trimmed = list(row)
    while trimmed and _is_empty(trimmed[-1]):
        trimmed.pop()
    return trimmed


def _extract_rows(
    worksheet,
    bounds: tuple[int, int, int, int],
    merged_values: dict[tuple[int, int], Any],
    max_rows: int,
    max_cols: int,
) -> tuple[list[list[Any]], bool, bool]:
    min_row, min_col, max_row, max_col = bounds

    row_limit = max_row
    col_limit = max_col
    rows_truncated = False
    cols_truncated = False

    if max_rows > 0 and (max_row - min_row + 1) > max_rows:
        row_limit = min_row + max_rows - 1
        rows_truncated = True
    if max_cols > 0 and (max_col - min_col + 1) > max_cols:
        col_limit = min_col + max_cols - 1
        cols_truncated = True

    rows: list[list[Any]] = []
    width = 0
    for row_index in range(min_row, row_limit + 1):
        row = [
            _cell_value(worksheet, row_index, col_index, merged_values)
            for col_index in range(min_col, col_limit + 1)
        ]
        row = _trim_trailing_empty_cells(row)
        width = max(width, len(row))
        rows.append(row)

    if width == 0:
        return [], rows_truncated, cols_truncated

    normalized_rows = [row + [""] * (width - len(row)) for row in rows]
    return normalized_rows, rows_truncated, cols_truncated


def _column_alignments(rows: list[list[Any]]) -> list[str]:
    if not rows:
        return []

    width = len(rows[0])
    alignments: list[str] = []
    data_rows = rows[1:] if len(rows) > 1 else rows
    for col_index in range(width):
        values = [row[col_index] for row in data_rows if not _is_empty(row[col_index])]
        if values and all(_is_numeric_value(value) for value in values):
            alignments.append("---:")
        else:
            alignments.append("---")
    return alignments


def _rows_to_markdown_table(rows: list[list[Any]]) -> str:
    if not rows:
        return "_No tabular content found._"

    formatted_rows = [[_format_cell_value(value) for value in row] for row in rows]
    width = len(formatted_rows[0])
    separator = _column_alignments(rows)
    lines = [
        "| " + " | ".join(formatted_rows[0]) + " |",
        "| " + " | ".join(separator or ["---"] * width) + " |",
    ]

    for row in formatted_rows[1:]:
        lines.append("| " + " | ".join(row) + " |")

    return "\n".join(lines)


# ─────────────────────────────────────────────────────────────
# Excel → Markdown
# ─────────────────────────────────────────────────────────────

def _convert_excel(input_file: Path, out_file: Path, max_rows: int, max_cols: int) -> str:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] openpyxl not installed. Run: pip install openpyxl")
        return ""

    workbook = load_workbook(input_file, data_only=True, read_only=False)
    visible_sheets = [sheet for sheet in workbook.worksheets if sheet.sheet_state == "visible"]

    lines: list[str] = [
        f"# Spreadsheet Source: {input_file.name}",
        "",
        "## Workbook Summary",
        "",
        f"- Sheets: {len(workbook.worksheets)}",
        f"- Visible sheets: {', '.join(sheet.title for sheet in visible_sheets) or 'None'}",
        "",
        "> Note: Formula cells are exported as cached values. This converter does not recalculate formulas.",
        "",
    ]

    if not visible_sheets:
        lines.extend(["_No visible sheets found._", ""])

    for worksheet in visible_sheets:
        merged_values = _merged_value_map(worksheet)
        bounds = _content_bounds(worksheet, merged_values)

        lines.extend([
            f"## Sheet: {worksheet.title}",
            "",
            f"- State: {_sheet_state_label(worksheet.sheet_state)}",
        ])

        if bounds is None:
            lines.extend(["", "_No content found._", ""])
            continue

        min_row, min_col, max_row, max_col = bounds
        used_range = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
        rows, rows_truncated, cols_truncated = _extract_rows(
            worksheet,
            bounds,
            merged_values,
            max_rows=max_rows,
            max_cols=max_cols,
        )

        lines.extend([
            f"- Used range: {used_range}",
            f"- Rows: {max_row - min_row + 1}",
            f"- Columns: {max_col - min_col + 1}",
            "",
        ])

        if rows_truncated or cols_truncated:
            limit_notes = []
            if rows_truncated:
                limit_notes.append(f"rows limited to {max_rows}")
            if cols_truncated:
                limit_notes.append(f"columns limited to {max_cols}")
            lines.extend([f"> Truncated: {', '.join(limit_notes)}.", ""])

        lines.extend([_rows_to_markdown_table(rows), ""])

    markdown = "\n".join(lines).rstrip() + "\n"
    out_file.write_text(markdown, encoding="utf-8")
    _report_result(out_file)
    return markdown


# ─────────────────────────────────────────────────────────────
# Dispatcher
# ─────────────────────────────────────────────────────────────

def convert_to_markdown(
    input_path: str,
    output_path: str | None = None,
    max_rows: int = 0,
    max_cols: int = 0,
) -> str:
    input_file = Path(input_path)
    if not input_file.exists():
        print(f"[ERROR] File not found: {input_path}")
        return ""

    suffix = input_file.suffix.lower()
    if suffix in LEGACY_EXCEL_FORMATS:
        print("[ERROR] Unsupported legacy Excel format: .xls")
        print("   Please resave the workbook as .xlsx and run this converter again.")
        return ""
    if suffix not in EXCEL_FORMATS:
        supported = ", ".join(sorted(EXCEL_FORMATS))
        print(f"[ERROR] Unsupported format: {suffix}")
        print(f"   Supported: {supported}")
        return ""

    if max_rows < 0 or max_cols < 0:
        print("[ERROR] --max-rows and --max-cols must be zero or positive integers")
        return ""

    out_file = Path(output_path) if output_path else input_file.with_suffix(".md")
    out_file.parent.mkdir(parents=True, exist_ok=True)

    print(f"[INFO] Converting Excel workbook: {input_file.name}")
    return _convert_excel(input_file, out_file, max_rows=max_rows, max_cols=max_cols)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert Excel workbooks to Markdown",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_to_md.py report.xlsx
  python excel_to_md.py report.xlsx -o output.md
  python excel_to_md.py report.xlsm --max-rows 200 --max-cols 40

Supported formats:
  .xlsx  .xlsm

Unsupported by default:
  .xls   Resave as .xlsx first
        """,
    )
    parser.add_argument("input", help="Input Excel workbook")
    parser.add_argument("-o", "--output", help="Output Markdown file path")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=0,
        help="Maximum rows per sheet to export (0 = no limit)",
    )
    parser.add_argument(
        "--max-cols",
        type=int,
        default=0,
        help="Maximum columns per sheet to export (0 = no limit)",
    )
    args = parser.parse_args()

    result = convert_to_markdown(
        args.input,
        args.output,
        max_rows=args.max_rows,
        max_cols=args.max_cols,
    )
    sys.exit(0 if result else 1)


if __name__ == "__main__":
    main()
