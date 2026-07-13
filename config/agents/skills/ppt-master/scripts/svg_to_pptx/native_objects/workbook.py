"""Minimal workbook generation for native chart data."""

from __future__ import annotations

import io
import zipfile
from typing import Any

try:
    from xlsxwriter import Workbook as XlsxWriterWorkbook
except ImportError:  # pragma: no cover - optional compatibility enhancement
    XlsxWriterWorkbook = None

try:
    from openpyxl import Workbook as OpenpyxlWorkbook
except ImportError:  # pragma: no cover - optional compatibility enhancement
    OpenpyxlWorkbook = None

from ..drawingml.utils import _xml_escape
from .marker_common import _excel_col


def _xlsx_cell_ref(row: int, col: int) -> str:
    return f"{_excel_col(col)}{row}"


def _xlsx_cell(value: Any, row: int, col: int) -> str:
    if value is None:
        return ""
    ref = _xlsx_cell_ref(row, col)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{ref}"><v>{value}</v></c>'
    return (
        f'<c r="{ref}" t="inlineStr"><is><t>{_xml_escape(str(value))}</t></is></c>'
    )


def _minimal_workbook(rows: list[list[Any]]) -> bytes:
    if XlsxWriterWorkbook is not None:
        buffer = io.BytesIO()
        workbook = XlsxWriterWorkbook(buffer, {
            "in_memory": True,
            "strings_to_formulas": False,
            "strings_to_urls": False,
        })
        worksheet = workbook.add_worksheet("Sheet1")
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                worksheet.write(row_index, col_index, value)
        workbook.close()
        return buffer.getvalue()

    if OpenpyxlWorkbook is not None:
        workbook = OpenpyxlWorkbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        for row_index, row in enumerate(rows, start=1):
            for col_index, value in enumerate(row, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=value)
                if isinstance(value, str):
                    cell.data_type = "s"
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    sheet_rows = []
    for row_index, values in enumerate(rows, start=1):
        cells = "".join(
            _xlsx_cell(value, row_index, col_index)
            for col_index, value in enumerate(values, start=1)
        )
        sheet_rows.append(f'<row r="{row_index}">{cells}</row>')

    entries = {
        "[Content_Types].xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml"
          ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
<Override PartName="/xl/worksheets/sheet1.xml"
          ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
<Override PartName="/xl/styles.xml"
          ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>''',
        "_rels/.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1"
              Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument"
              Target="xl/workbook.xml"/>
</Relationships>''',
        "xl/workbook.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
          xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets>
</workbook>''',
        "xl/_rels/workbook.xml.rels": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1"
              Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
              Target="worksheets/sheet1.xml"/>
<Relationship Id="rId2"
              Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles"
              Target="styles.xml"/>
</Relationships>''',
        "xl/worksheets/sheet1.xml": f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
           xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheetData>{''.join(sheet_rows)}</sheetData>
</worksheet>''',
        "xl/styles.xml": '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>
<fills count="1"><fill><patternFill patternType="none"/></fill></fills>
<borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
</styleSheet>''',
    }

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in entries.items():
            zout.writestr(name, data.encode("utf-8"))
    return buffer.getvalue()


def _minimal_category_chart_workbook(chart_data: dict[str, Any]) -> bytes:
    if chart_data.get("kind") == "combo" and chart_data.get("independent_categories"):
        plots = chart_data["plots"]
        column_count = max(
            int(plot["start_column"]) + len(plot["series"]) - 1
            for plot in plots
        )
        rows: list[list[Any]] = [[None] * column_count]
        for plot in plots:
            category_column = int(plot["category_column"]) - 1
            start_column = int(plot["start_column"]) - 1
            for offset, series in enumerate(plot["series"]):
                rows[0][start_column + offset] = series["name"]
            for point_index, category in enumerate(plot["categories"], start=1):
                while len(rows) <= point_index:
                    rows.append([None] * column_count)
                rows[point_index][category_column] = category
                for offset, series in enumerate(plot["series"]):
                    rows[point_index][start_column + offset] = (
                        series["values"][point_index - 1]
                    )
        return _minimal_workbook(rows)

    categories = chart_data["categories"]
    series = chart_data["series"]
    rows: list[list[Any]] = [[None] + [item["name"] for item in series]]
    for row_index, category in enumerate(categories):
        rows.append([category] + [item["values"][row_index] for item in series])
    return _minimal_workbook(rows)


def _minimal_xy_chart_workbook(chart_data: dict[str, Any]) -> bytes:
    series = chart_data["series"]
    is_bubble = chart_data["type"] == "bubble"
    rows: list[list[Any]] = [[]]
    for item in series:
        rows[0].extend([f"{item['name']} X", item["name"]])
        if is_bubble:
            rows[0].append(f"{item['name']} Size")

    max_points = max(len(item["x"]) for item in series)
    for point_idx in range(max_points):
        row: list[Any] = []
        for item in series:
            if point_idx < len(item["x"]):
                row.extend([item["x"][point_idx], item["y"][point_idx]])
                if is_bubble:
                    row.append(item["sizes"][point_idx])
            else:
                row.extend(["", ""])
                if is_bubble:
                    row.append("")
        rows.append(row)
    return _minimal_workbook(rows)


def _minimal_chart_ex_workbook(chart_data: dict[str, Any]) -> bytes:
    chart_type = chart_data["type"]
    if chart_type in {"sunburst", "treemap"}:
        levels = chart_data["levels"]
        rows: list[list[Any]] = [[f"Level {idx + 1}" for idx in range(len(levels))] + ["Value"]]
        for row_idx, value in enumerate(chart_data["values"]):
            rows.append([level[row_idx] for level in levels] + [value])
        return _minimal_workbook(rows)
    if chart_type == "histogram":
        return _minimal_workbook([["Value"]] + [[value] for value in chart_data["values"]])
    if chart_type in {"funnel", "pareto", "waterfall"}:
        rows = [["Category", "Value"]]
        rows.extend(
            [category, chart_data["values"][idx]]
            for idx, category in enumerate(chart_data["categories"])
        )
        return _minimal_workbook(rows)
    if chart_type == "box_whisker":
        rows = [[]]
        for item in chart_data["series"]:
            rows[0].extend([f"{item['name']} Category", item["name"]])
        max_rows = max(len(item["values"]) for item in chart_data["series"])
        for row_idx in range(max_rows):
            row: list[Any] = []
            for item in chart_data["series"]:
                if row_idx < len(item["values"]):
                    row.extend([item["categories"][row_idx], item["values"][row_idx]])
                else:
                    row.extend(["", ""])
            rows.append(row)
        return _minimal_workbook(rows)
    raise RuntimeError(f"Native PPTX {chart_type} chart is outside current basic chart support")
